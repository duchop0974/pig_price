"""Xuất dữ liệu ra Excel: giá heo hơi và kế hoạch xuất bán."""
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from core.repositories.prices_repo import load_records_df
from core.repositories.sale_orders_repo import list_orders_for_export
from core.repositories.sale_plans_repo import list_sale_plans_for_export

EXPORT_COLUMNS = {
    "date": "Ngày",
    "source": "Nguồn",
    "region": "Miền",
    "province": "Địa phương",
    "price_vnd_per_kg": "Giá (đ/kg)",
    "change_vnd_per_kg": "Biến động (đ/kg)",
    "benchmark_price_vnd_per_kg": "Giá heo cám GreenFeed (đ/kg)",
    "source_url": "Nguồn URL",
}


def export_to_excel(db_path: Path, dest) -> int:
    """Xuất toàn bộ dữ liệu ra Excel. `dest` có thể là đường dẫn file hoặc
    buffer (BytesIO, dùng khi trả file trực tiếp qua web). Trả về số dòng
    đã xuất."""
    df = load_records_df(db_path)
    if df.empty:
        raise ValueError("Chưa có dữ liệu để xuất.")

    export_df = df.copy()
    export_df["_date_sort"] = pd.to_datetime(export_df["date"], format="%d/%m/%Y")
    export_df = export_df.sort_values(["_date_sort", "source", "province"]).drop(columns="_date_sort")
    export_df = export_df.rename(columns=EXPORT_COLUMNS)

    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Gia heo hoi")
        ws = writer.sheets["Gia heo hoi"]
        for col_idx, col_name in enumerate(export_df.columns, start=1):
            max_len = max(
                len(str(col_name)),
                int(export_df.iloc[:, col_idx - 1].fillna("").astype(str).str.len().max())
                if len(export_df)
                else 0,
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 45)
        ws.freeze_panes = "A2"

    return len(export_df)


SALE_PLAN_EXPORT_COLUMNS = {
    "plan_code": "Mã kế hoạch",
    "id": "ID nội bộ",
    "planned_date": "Ngày dự kiến",
    "farm": "Trang trại",
    "province": "Tỉnh",
    "zone": "Khu",
    "shed": "Chuồng",
    "lot": "Lô",
    "pig_type_name": "Loại heo",
    "quantity": "Số lượng dự kiến (con)",
    "received_quantity": "Số lượng thực nhận (con)",
    "received_at": "Ngày nhận",
    "received_by": "Người nhận",
    "allocated_quantity": "Số lượng đã phân bổ (con)",
    "remaining_quantity": "Số lượng còn lại (con)",
    "actual_sold_quantity": "Đã xuất/giao thực tế (con)",
    "reconciled_quantity": "Đã đối soát (con)",
    "remaining_to_reconcile": "Chưa xử lý (con)",
    "expected_avg_weight_kg": "Trọng lượng dự kiến (kg/con)",
    "planned_total_weight_kg": "Khối lượng kế hoạch (kg)",
    "actual_total_weight_kg": "Khối lượng thực tế (kg)",
    "weight_missing_delivery_count": "Số lần xuất chưa nhập cân",
    "delivery_count": "Số lần xuất",
    "last_delivered_date": "Ngày xuất gần nhất",
    "off_type_quantity": "Số con lệch loại (con)",
    "zone_id": "ID khu (nội bộ)",
    "pig_type_id": "ID loại heo (nội bộ)",
    "locked_at": "Khoá lúc",
    "locked_by": "Khoá bởi",
    "note": "Ghi chú",
    "status": "Trạng thái",
    "created_by": "Người tạo",
    "approved_by": "Duyệt bởi",
    "approved_at": "Duyệt lúc",
    "rejected_by": "Từ chối bởi",
    "rejected_at": "Từ chối lúc",
    "rejected_reason": "Lý do từ chối",
    "created_at": "Tạo lúc",
    "created_ip": "IP tạo",
    "updated_by": "Người sửa gần nhất",
    "updated_at": "Sửa lúc",
    "updated_ip": "IP sửa",
}

SALE_ORDER_EXPORT_COLUMNS = {
    "order_code": "Mã đơn hàng",
    "order_id": "ID đơn (nội bộ)",
    "plan_code": "Mã dòng hàng",
    "id": "ID dòng (nội bộ)",
    "sale_plan_code": "Mã kế hoạch trại",
    "planned_date": "Ngày dự kiến",
    "farm": "Trang trại",
    "province": "Tỉnh",
    "zone": "Khu",
    "shed": "Chuồng",
    "lot": "Lô",
    "pig_type_name": "Loại heo",
    "quantity": "Số lượng (con)",
    "selling_price": "Giá chào bán (đ/kg)",
    "note": "Ghi chú dòng",
    "actual_price": "Giá bán thực tế (đ/kg)",
    "actual_quantity": "Số lượng bán thực tế (con)",
    "customer_name": "Khách hàng",
    "customer_phone": "SĐT khách hàng",
    "customer_email": "Email khách hàng",
    "customer_contact_person": "Người liên hệ khách hàng",
    "contact_note": "Ghi chú liên hệ",
    "contacted_by": "Người liên hệ",
    "contacted_at": "Liên hệ lúc",
    "confirmed_sale_at": "Ngày chốt bán",
    "delivery_time": "Khung giờ giao",
    "payment_method": "Hình thức thanh toán",
    "paid_amount": "Số tiền đã thu (đ)",
    "paid_at": "Ngày thu tiền",
    "weighing_ref": "Số chứng từ cân",
    "invoice_number": "Số hoá đơn",
    "invoiced_by": "Người lập hoá đơn",
    "invoiced_at": "Lập hoá đơn lúc",
    "revenue_recorded_by": "Người ghi nhận doanh thu",
    "revenue_recorded_at": "Ghi nhận doanh thu lúc",
    "status": "Trạng thái đơn",
    "created_by": "Người tạo",
    "created_at": "Tạo lúc",
    "created_ip": "IP tạo",
    "updated_by": "Người sửa gần nhất",
    "updated_at": "Sửa lúc",
    "updated_ip": "IP sửa",
}

# Thứ tự cột cho bảng khi đơn có >=2 dòng hàng (STT thêm ở runtime) — khớp
# đúng thứ tự đọc tự nhiên "Loại heo | Trang trại | Số lượng | Đơn giá |
# Ngày dự kiến". Không có order_code (hiện 1 lần ở khối tiêu đề, không lặp
# theo dòng), không có shed/lot (thông tin quản lý trại nội bộ), không có
# delivery_time/payment_method (field cấp ĐƠN, hiện riêng trong "Thông tin
# bổ sung" — xem export_order_quotation_to_excel).
QUOTATION_TABLE_FIELDS = [
    ("pig_type_name", "Loại heo"),
    ("farm", "Trang trại"),
    ("quantity", "Số lượng"),
    ("selling_price", "Đơn giá"),
    ("planned_date", "Ngày dự kiến"),
]

PAYMENT_METHOD_LABEL = {
    "bank_transfer_immediate": "Chuyển khoản ngay",
    "bank_transfer_24h": "Chuyển khoản trước 24h",
    "cash": "Tiền mặt",
    "credit": "Công nợ",
    "other": "Khác",
}

SALE_PLAN_STATUS_LABEL = {
    "active": "Đang chờ",  # fallback cho dữ liệu cũ, không còn ghi mới giá trị này
    "pending_approval": "Chờ duyệt",
    "approved": "Đã duyệt",
    "rejected": "Từ chối",
    "cancelled": "Đã hủy",
    "disabled": "Đã vô hiệu hoá",
}

ORDER_STATUS_LABEL = {
    "active": "Đang xử lý",
    "done": "Đã bán",
    "cancelled": "Đã hủy",
    "disabled": "Đã vô hiệu hoá",
}


def _autosize_and_freeze(df: pd.DataFrame, ws) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            int(df.iloc[:, col_idx - 1].fillna("").astype(str).str.len().max()) if len(df) else 0,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 45)
    ws.freeze_panes = "A2"


def export_sale_plans_to_excel(db_path: Path, dest) -> int:
    """Xuất toàn bộ kế hoạch trại (nguồn cung, BM01) ra Excel, gồm cả trường
    ẩn để đối soát. `dest` có thể là đường dẫn file hoặc buffer (BytesIO)."""
    rows = list_sale_plans_for_export(db_path)
    if not rows:
        raise ValueError("Chưa có kế hoạch nào để xuất.")

    df = pd.DataFrame(rows).drop(columns=["pig_type", "farm_id"])
    df["status"] = df["status"].map(lambda s: SALE_PLAN_STATUS_LABEL.get(s, s))
    df = df.rename(columns=SALE_PLAN_EXPORT_COLUMNS)

    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ke hoach trai")
        _autosize_and_freeze(df, writer.sheets["Ke hoach trai"])

    return len(df)


def export_sale_orders_to_excel(db_path: Path, dest) -> int:
    """Xuất toàn bộ đơn hàng (Phòng bán hàng, BM02, Giai đoạn 8) ra Excel, gồm
    cả trường ẩn để đối soát — 1 dòng Excel = 1 dòng hàng (đơn nhiều dòng lặp
    lại đúng số dòng, cùng order_code)."""
    rows = list_orders_for_export(db_path)
    if not rows:
        raise ValueError("Chưa có đơn hàng nào để xuất.")

    df = pd.DataFrame(rows).drop(columns=["pig_type", "sale_plan_id", "customer_id"])
    df["status"] = df["status"].map(lambda s: ORDER_STATUS_LABEL.get(s, s))
    df["payment_method"] = df["payment_method"].map(lambda s: PAYMENT_METHOD_LABEL.get(s, s))
    df = df.rename(columns=SALE_ORDER_EXPORT_COLUMNS)

    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ke hoach ban")
        _autosize_and_freeze(df, writer.sheets["Ke hoach ban"])

    return len(df)


# Màu lấy đúng từ design token webapp/static/css/style.css (:root) — Excel
# không đọc được CSS variable nên quy đổi thẳng sang mã hex tương ứng.
# Tiết chế: chỉ dùng làm accent (chữ), không fill nền đậm cho mọi ô như
# bảng database.
_HEX_PRIMARY_DARK = "00507F"
_HEX_TEXT = "1C2430"
_HEX_MUTED = "667085"
_HEX_BORDER = "E3E7EE"
_HEX_SUMMARY_FILL = "F3F7FA"  # nền rất nhạt cho khối tóm tắt thương mại 3 cột

# Logo có sẵn của dự án (webapp/static/img/) — không tự tạo logo/thông tin
# công ty giả, chỉ dùng asset đã tồn tại.
_LOGO_PATH = Path(__file__).resolve().parent.parent.parent / "webapp" / "static" / "img" / "logo-icon.png"

_QUOTATION_CANVAS_COLS = 7  # canvas nội dung cố định A..G, không kéo rộng theo dữ liệu
_QUOTATION_COL_WIDTHS = {"A": 16, "B": 2, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14}


def _thin_border() -> Border:
    side = Side(style="thin", color=_HEX_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _set_row_border(ws, row: int, col_start: int, col_end: int, *, top=False, bottom=False) -> None:
    """Kẻ 1 đường mảnh ngang (không có chữ) trên/dưới 1 dòng — dùng làm
    ranh giới giữa các khối thay vì border bao quanh từng ô kiểu bảng."""
    side = Side(style="thin", color=_HEX_BORDER)
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(top=side if top else None, bottom=side if bottom else None)


def _set_outer_box(ws, row_start: int, row_end: int, col_start: int, col_end: int, *, fill: str | None = None) -> None:
    """Viền mảnh CHỈ ở cạnh ngoài của 1 khối ô chữ nhật (không kẻ lưới bên
    trong) — dùng cho khối tóm tắt thương mại, tránh trông như 1 bảng dữ
    liệu."""
    side = Side(style="thin", color=_HEX_BORDER)
    pattern = PatternFill(fill_type="solid", fgColor=fill) if fill else None
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=side if c == col_start else None,
                right=side if c == col_end else None,
                top=side if r == row_start else None,
                bottom=side if r == row_end else None,
            )
            if pattern:
                cell.fill = pattern


def _fmt_vi_number(n) -> str:
    """1000 -> '1.000' (dấu chấm ngăn cách hàng nghìn kiểu Việt Nam, khớp
    fmtPrice()/toLocaleString("vi-VN") đang dùng ở webapp/static/js)."""
    return f"{n:,.0f}".replace(",", ".")


def _fmt_date_ddmmyyyy(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        y, m, d = iso[:10].split("-")
        return f"{d}/{m}/{y}"
    except ValueError:
        return iso


def _quotation_line_value(row: dict, key: str) -> str:
    """Giá trị hiển thị cho 1 field cấp dòng hàng — field bắt buộc (luôn có
    dữ liệu thật theo ràng buộc tạo đơn) mà rỗng thì để trống ô, KHÔNG hiện
    text kiểu 'Chưa cập nhật'/'N/A' (tài liệu gửi khách, không phải màn
    hình quản trị)."""
    if key == "planned_date":
        return _fmt_date_ddmmyyyy(row.get(key))
    if key == "quantity":
        qty = row.get(key)
        return f"{_fmt_vi_number(qty)} con" if qty is not None else ""
    if key == "selling_price":
        price = row.get(key)
        return f"{_fmt_vi_number(price)} đ/kg" if price else ""
    return row.get(key) or ""


def export_order_quotation_to_excel(db_path: Path, dest, order_ids: list[int]) -> int:
    """Xuất 'Bảng chào giá' — tài liệu thương mại gửi trực tiếp cho khách
    hàng, bố cục kiểu commercial offer (không phải bảng dữ liệu hệ thống):
    1 khối tiêu đề thống nhất (logo+title+mã/ngày), lời chào, khối tóm tắt
    thương mại nổi bật (Loại heo/Số lượng/Đơn giá — layout phiếu khi 1
    dòng hàng, bảng khi nhiều dòng), section "Thông tin bổ sung" chỉ hiện
    field có dữ liệu, và lời kết. Canvas nội dung cố định 7 cột (A-G),
    không kẻ gridline. Field trống/không bắt buộc (khách hàng, thanh
    toán, khung giờ giao, ghi chú) bị bỏ hẳn khỏi tài liệu khi không có
    dữ liệu — không hiện "Chưa cập nhật"/"N/A". Không có cột ID/audit/kỹ
    thuật, không có chuồng/lô (thông tin quản lý trại nội bộ), và KHÔNG
    tính "Tổng tiền" — hệ thống chỉ có số lượng theo con và giá theo kg,
    không có tổng trọng lượng nên nhân trực tiếp sẽ ra số sai đơn vị."""
    rows = [r for r in list_orders_for_export(db_path) if r["order_id"] in set(order_ids)]
    if not rows:
        raise ValueError("Không tìm thấy đơn hàng để xuất.")
    for r in rows:
        r["payment_method"] = PAYMENT_METHOD_LABEL.get(r["payment_method"], r["payment_method"])

    last_col = _QUOTATION_CANVAS_COLS
    last_col_letter = get_column_letter(last_col)
    single_order = len(order_ids) == 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Bang chao gia"
    ws.sheet_view.showGridLines = False
    for col, width in _QUOTATION_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # --- Khối tiêu đề thống nhất: logo + title + subtitle + mã/ngày ---
    if _LOGO_PATH.exists():
        img = XLImage(str(_LOGO_PATH))
        img.width = 44
        img.height = 44
        ws.add_image(img, "A1")
    ws.merge_cells(f"C1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=3, value="BẢNG CHÀO GIÁ")
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=_HEX_PRIMARY_DARK)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C2:{last_col_letter}2")
    sub_cell = ws.cell(row=2, column=3, value="HEO THƯƠNG PHẨM")
    sub_cell.font = Font(name="Calibri", size=12, color=_HEX_MUTED)
    sub_cell.alignment = Alignment(horizontal="center")

    o0 = rows[0]
    meta_text = f"Mã chào giá: {o0.get('order_code') or ''}   ·   Ngày: {datetime.now():%d/%m/%Y}" if single_order else f"Ngày: {datetime.now():%d/%m/%Y}"
    ws.merge_cells(f"C3:{last_col_letter}3")
    meta_cell = ws.cell(row=3, column=3, value=meta_text)
    meta_cell.font = Font(name="Calibri", size=10, color=_HEX_MUTED)
    meta_cell.alignment = Alignment(horizontal="center")

    _set_row_border(ws, 4, 1, last_col, bottom=True)
    row = 6

    # --- Lời chào ---
    ws.cell(row=row, column=1, value="Kính gửi Quý khách,").font = Font(name="Calibri", color=_HEX_TEXT)
    row += 1
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    intro_cell = ws.cell(
        row=row,
        column=1,
        value="Xuân Thiện trân trọng gửi Quý khách thông tin chào giá heo thương phẩm theo nội dung dưới đây:",
    )
    intro_cell.font = Font(name="Calibri", color=_HEX_TEXT)
    intro_cell.alignment = Alignment(wrap_text=True)
    row += 2

    # --- THÔNG TIN CHÀO GIÁ: khối tóm tắt nổi bật (1 dòng hàng) hoặc bảng (nhiều dòng) ---
    section_row = row
    section_cell = ws.cell(row=row, column=1, value="THÔNG TIN CHÀO GIÁ")
    section_cell.font = Font(name="Calibri", size=12, bold=True, color=_HEX_PRIMARY_DARK)
    _set_row_border(ws, section_row, 1, last_col, bottom=True)
    row += 2

    if len(rows) == 1:
        o = rows[0]
        summary_fields = [
            ("pig_type_name", "LOẠI HEO", 1, 2),
            ("quantity", "SỐ LƯỢNG", 3, 4),
            ("selling_price", "ĐƠN GIÁ", 5, 7),
        ]
        label_row, value_row = row, row + 1
        for key, label, col_start, col_end in summary_fields:
            start_letter, end_letter = get_column_letter(col_start), get_column_letter(col_end)
            ws.merge_cells(f"{start_letter}{label_row}:{end_letter}{label_row}")
            lbl = ws.cell(row=label_row, column=col_start, value=label)
            lbl.font = Font(name="Calibri", size=9, bold=True, color=_HEX_MUTED)
            lbl.alignment = Alignment(horizontal="center")
            ws.merge_cells(f"{start_letter}{value_row}:{end_letter}{value_row}")
            val = ws.cell(row=value_row, column=col_start, value=_quotation_line_value(o, key))
            val.font = Font(name="Calibri", size=16, bold=True, color=_HEX_PRIMARY_DARK)
            val.alignment = Alignment(horizontal="center")
        ws.row_dimensions[value_row].height = 26
        _set_outer_box(ws, label_row, value_row, 1, last_col, fill=_HEX_SUMMARY_FILL)
        row = value_row + 2

        # --- THÔNG TIN BỔ SUNG: field phụ (luôn có Trang trại/Ngày dự kiến, thêm field cấp đơn nếu có) ---
        extra_fields = [
            (label, value)
            for label, value in (
                ("Trang trại", o.get("farm")),
                ("Ngày dự kiến", _fmt_date_ddmmyyyy(o.get("planned_date"))),
                ("Khách hàng", o.get("customer_name")),
                ("Hình thức thanh toán", o.get("payment_method")),
                ("Khung giờ giao", o.get("delivery_time")),
                ("Ghi chú", o.get("note")),
            )
            if value
        ]
        if extra_fields:
            extra_section_row = row
            extra_cell = ws.cell(row=row, column=1, value="THÔNG TIN BỔ SUNG")
            extra_cell.font = Font(name="Calibri", size=10, bold=True, color=_HEX_MUTED)
            _set_row_border(ws, extra_section_row, 1, last_col, bottom=True)
            row += 1
            for label, value in extra_fields:
                label_cell = ws.cell(row=row, column=1, value=label)
                label_cell.font = Font(name="Calibri", size=10, bold=True, color=_HEX_TEXT)
                ws.merge_cells(f"C{row}:{last_col_letter}{row}")
                value_cell = ws.cell(row=row, column=3, value=value)
                value_cell.font = Font(name="Calibri", size=10, color=_HEX_TEXT)
                row += 1
        last_data_row = row - 1
    else:
        show_note_col = any(o.get("note") for o in rows)
        table_fields = [("_stt", "STT")] + list(QUOTATION_TABLE_FIELDS) + ([("note", "Ghi chú")] if show_note_col else [])
        thin = _thin_border()
        header_row = row
        for col, (_key, label) in enumerate(table_fields, start=1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = Font(name="Calibri", size=10, bold=True, color=_HEX_PRIMARY_DARK)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
        row += 1
        for idx, o in enumerate(rows, start=1):
            for col, (key, _label) in enumerate(table_fields, start=1):
                value = idx if key == "_stt" else _quotation_line_value(o, key)
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center" if key == "_stt" else "left", wrap_text=True)
                cell.font = Font(name="Calibri", size=10)
            row += 1
        last_data_row = row - 1
        # Bảng dùng cột KHÔNG merge (mỗi field 1 cột riêng) — độ rộng mặc
        # định của canvas (vd cột B=2, dành cho spacer ở layout phiếu 1
        # dòng) không phù hợp, phải tính lại theo nội dung thật của bảng.
        for col_idx, (_key, label) in enumerate(table_fields, start=1):
            cell_lens = (len(str(ws.cell(r, col_idx).value or "")) for r in range(header_row, last_data_row + 1))
            max_len = max(len(label), *cell_lens)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 6), 30)
        row += 1

        # --- THÔNG TIN BỔ SUNG: chỉ field cấp-đơn (Trang trại/Ngày dự kiến đã có trong bảng) ---
        o = rows[0]
        extra_fields = [
            (label, value)
            for label, value in (
                ("Khách hàng", o.get("customer_name")),
                ("Hình thức thanh toán", o.get("payment_method")),
                ("Khung giờ giao", o.get("delivery_time")),
            )
            if value
        ] if single_order else []
        if extra_fields:
            extra_section_row = row
            extra_cell = ws.cell(row=row, column=1, value="THÔNG TIN BỔ SUNG")
            extra_cell.font = Font(name="Calibri", size=10, bold=True, color=_HEX_MUTED)
            _set_row_border(ws, extra_section_row, 1, last_col, bottom=True)
            row += 1
            for label, value in extra_fields:
                label_cell = ws.cell(row=row, column=1, value=label)
                label_cell.font = Font(name="Calibri", size=10, bold=True, color=_HEX_TEXT)
                ws.merge_cells(f"C{row}:{last_col_letter}{row}")
                value_cell = ws.cell(row=row, column=3, value=value)
                value_cell.font = Font(name="Calibri", size=10, color=_HEX_TEXT)
                row += 1
            last_data_row = row - 1

    # --- Lời kết ---
    row = last_data_row + 1
    _set_row_border(ws, row, 1, last_col, top=True)
    row += 2
    ws.cell(row=row, column=1, value="Trân trọng,").font = Font(name="Calibri", italic=True, color=_HEX_MUTED)
    row += 1
    ws.cell(row=row, column=1, value="XUÂN THIỆN").font = Font(name="Calibri", size=12, bold=True, color=_HEX_PRIMARY_DARK)
    last_data_row = row

    # --- Page setup: A4 dọc, canvas hẹp 7 cột, fit 1 trang ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.print_area = f"A1:{last_col_letter}{last_data_row}"

    wb.save(dest)
    return len(rows)
